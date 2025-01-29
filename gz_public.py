import random
import os
import re
import subprocess

import openpyxl
import psutil
import yaml
import initPhone
import requests
import logging
import json
import uuid
from hashlib import md5
import urllib3
import time
import datetime
from datetime import datetime, date
import allure
import pytesseract
from PIL import Image, ImageOps
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

# 远程Appium服务地址
# gzAppiumH = "http://10.10.1.123:4723/wd/hub"
gzAppiumH = "http://127.0.0.1:4723/wd/hub"

# gzHostCnTmp = 'api-cn.snser.wang'
gzHostCnTmp = 'api-cn.aosulife.com'

# 手机系统
gzPlatformN = "Android"

# 包信息
gzAppPack = initPhone.get_package_name()
gzAppActivity = "com.glazero.android.SplashActivity"

# 登录账号
# email = "1499405887@qq.com"
# pwd = "Qwe222222"

# email = "17332359401@163.com"
# pwd = "Aa123456"

# email = "2802662060@qq.com"
# pwd = "2802662060@Ch"

email = "liushijie0041@163.com"
pwd = "Aa123456"

# home_user = 'enoch@glazero.com'
home_user = 'liushijie0050@163.com'
home_user_pwd = 'Qwe123456'

# 修改后的密码
change_pwd_to = 'Qwe101010'

# 业务类型，默认是1
_type = 1

# 手机列表
gzPhoneList = [
    # {'gzDeviceMode': "samsungA51", 'gzDeviceN': "10.10.1.157:5555", 'gzPlatformVer': "10"},
    {'gzDeviceMode': "samsungA51", 'gzDeviceN': "R58N828LKMP", 'gzPlatformVer': "10"},
    {'gzDeviceMode': "samsungS10e", 'gzDeviceN': "10.10.1.20:5555", 'gzPlatformVer': "11"}
]

# 业务参数
gzRegionList = [
    {'CN': "中国区", 'region': "中国", 'code': "+86"},
    {'US': "美国区", 'region': "美国", 'code': "+1"}
]

# default 默认连接三星A51 中国区
# CONNECTTO = gzPhoneList[0]['gzDeviceN']
CONNECTTO = initPhone.get_dev_id()
# PLATFORMVER = gzPhoneList[0]['gzPlatformVer']
PLATFORMVER = initPhone.get_android_version()
REGION = gzRegionList[0]['code']


# 生成随机邮箱
def randomEmail():
    metaData = "0123456789abcdefghijklmnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ"
    emailType = ['@qq.com', '@163.com', '@outlook.com', '@glazero.com']
    emailLength = random.randint(4, 10)
    emailPrefox = "".join(random.choice(metaData) for i in range(emailLength))
    emailSuffix = random.choice(emailType)
    email = emailPrefox + emailSuffix
    return email


# 判断元素是否存在
def isElementPresent(driver, by, value):
    try:
        driver.find_element(by=by, value=value)
    except Exception as e:
        # 打印异常信息
        print(e)
        # 发生异常，表示页面中没有该元素
        return False
    else:
        # 没有发生异常，表示页面存在该元素
        return True


# 获取屏幕尺寸
def get_page_size(driver):
    x = driver.get_window_size()['width']
    y = driver.get_window_size()['height']
    return x, y


# 下拉刷新
def swipe_down(driver):
    s = get_page_size(driver)
    sx = s[0] * 0.50
    sy = s[1] * 0.33
    ex = s[0] * 0.50
    ey = s[1] * 0.75
    driver.swipe(sx, sy, ex, ey)


def _md5(_pwd):
    password_encode = _pwd.encode('utf-8')
    password_md5 = md5(password_encode)
    password_hex = password_md5.hexdigest()
    return password_hex


def _headers():
    headers = {}
    headers['Content-Type'] = 'application/x-www-form-urlencoded'
    headers['User-Agent'] = 'PostmanRuntime/7.29.0'
    headers['Accept'] = '*/*'
    headers['Accept-Encoding'] = 'gzip, deflate, br'
    headers['Connection'] = 'keep-alive'
    headers['Gz-Pid'] = 'glazero'
    headers['Gz-Brand'] = 'samsung'
    headers['Imei'] = 'd71c95c2ea6bd816'
    headers['Gz-AppId'] = 'com.glazero.android'
    headers['Gz-AppVer'] = '1.3.0.2409'
    headers['Gz-Channel'] = 'internal'
    headers['Gz-FontSize'] = '1.1'
    headers['Gz-Imei'] = 'd71c95c2ea6bd816'
    headers['Gz-Lang'] = 'zh-Hans-CN'
    headers['Gz-Model'] = 'SM-A515U'
    headers[
        'Gz-NotifyPermission'] = '{filter=ALL, channels={%E9%97%A8%E9%93%83%E8%A2%AB%E5%BC%BA%E6%8B%86%E6%B6%88%E6%81%AF=4, %E6%8C%89%E9%97%A8%E9%93%83%E6%B6%88%E6%81%AF=4, %E4%BD%8E%E7%94%B5%E9%87%8F%E6%B6%88%E6%81%AF=4, %E9%97%A8%E9%93%83%E4%BA%8B%E4%BB%B6%E6%B6%88%E6%81%AF=4, %E5%85%B6%E4%BB%96%E6%9C%AA%E5%88%86%E7%BB%84%E6%B6%88%E6%81%AF=4, default=1}, channelGroups={default_group=false}, enabled=true, paused=false, BubblesAllowed=true}'
    headers['Gz-OsLang'] = ''
    headers['Gz-OsType'] = 'android'
    headers['Gz-OsVer'] = '29'
    headers['Gz-Sid'] = ''
    headers['Gz-Timezone'] = '+08:00'
    headers['Gz-Uid'] = ''
    return headers


def aosu_headers():
    headers = {'Content-Type': 'application/x-www-form-urlencoded'}
    return headers


# 登录接口，获取sessionId，为修改密码接口提供必要的header
def _login(gz_host, _email, _region, country_code, _password, _type=1):
    global SID, UID
    url = 'https://' + gz_host + '/v1/user/login' + '?' + 'uuid=' + 'android_ui_auto' + '&' + 't=' + '001'
    data = 'countryAbbr=' + _region + '&' + 'countryCode=' + country_code + '&' + 'email=' + _email + '&' + 'password=' + _password + '&' + 'type=%d' % _type
    rsp = requests.post(url, headers=_headers(), data=data, timeout=(10, 10), verify=False)
    # rsp_json = rsp.json()
    SID = rsp.json()['data']['sid']
    UID = rsp.json()['data']['uid']


def change_password(old_pwd, new_pwd, _email, _type, gz_host, _region='CN', country_code='86'):
    old_pwd_md5 = _md5(old_pwd)
    new_pwd_md5 = _md5(new_pwd)
    _login(gz_host, _email, _region, country_code, old_pwd_md5, _type)
    url = 'https://' + gz_host + '/v1/user/changePassword' + '?' + 'uuid=' + 'android_ui_auto' + '&' + 't=' + '002'
    data = 'email=' + _email + '&' + 'newPassword=' + new_pwd_md5 + '&' + 'oldPassword=' + old_pwd_md5 + '&' + 'type=%d' % _type
    headers = _headers()
    headers['Gz-Sid'] = SID
    headers['Gz-Uid'] = UID
    requests.post(url, headers=headers, data=data, timeout=(10, 10), verify=False)
    # rsp_json = rsp.json()


def _unbind(sn='V8P1AH110002353', dev_type=1, delete_cloud_data=0, gz_host=gzHostCnTmp):
    pwd_md5 = _md5(pwd)
    _login(gz_host, _email=email, _region='CN', country_code='86', _password=pwd_md5, _type=1)
    url = 'https://' + gz_host + '/v1/bind/unbind' + '?' + 'uuid=' + 'android_ui_auto' + '&' + 't=' + '003'
    data = 'sn=' + sn + '&' + 'devType=%d' % dev_type + '&' + 'deleteCloudData=%d' % delete_cloud_data
    headers = _headers()
    headers['Gz-Sid'] = SID
    headers['Gz-Uid'] = UID
    rsp = requests.post(url, headers=headers, data=data, timeout=(10, 10), verify=False)
    logging.info(rsp.json())
    if rsp.json() == {'errno': 0, 'errmsg': '成功', 'data': {}}:
        print('解绑成功：', rsp.json())
    elif rsp.json() == {'errno': 701, 'errmsg': '已解绑或者未绑定', 'data': {}}:
        print('已解绑：', rsp.json())


def get_device_name(model='V8P', gz_host=gzHostCnTmp):
    """
    获取设备列表中指定的model的设备，并将其返回
    :param gz_host: 域名，默认为中国区
    :param model: 默认是V8P
    :return: 将设备的名称返回
    """
    pwd_md5 = _md5(pwd)
    _login(gz_host, _email=email, _region='CN', country_code='86', _password=pwd_md5, _type=1)
    url = 'https://' + gz_host + '/v1/dev/getList' + '?' + 'uuid=' + 'android_ui_auto' + '&' + 't=' + '004'
    headers = _headers()
    headers['Gz-Sid'] = SID
    headers['Gz-Uid'] = UID
    rsp = requests.post(url, headers=headers, timeout=(10, 10), verify=False)
    logging.info(rsp.json())

    # 变量设备列表，筛选指定设备类型的设备名称并返回，例如，默认类型是V8P
    devices = list(rsp.json()["data"]["list"])
    if devices:
        for device in devices:
            # 获取指定设备类型的设备的名称，并且是主人设备，并且是在线的设备 满足条件的第一个
            if device["model"] == model and device["role"] == 0 and device["online"] == 1:
                dev_name = device["name"]
                return dev_name


def get_devices_list(gz_host=gzHostCnTmp):
    """
    获取设备列表中指定的model的设备，并将其返回
    :param gz_host: 域名，默认为中国区
    :param model: 默认是V8P
    :return: 将设备的名称返回
    """
    pwd_md5 = _md5(pwd)
    _login(gz_host, _email=email, _region='CN', country_code='86', _password=pwd_md5, _type=1)
    url = 'https://' + gz_host + '/v1/dev/getList' + '?' + 'uuid=' + 'android_ui_auto' + '&' + 't=' + '004'
    headers = _headers()
    headers['Gz-Sid'] = SID
    headers['Gz-Uid'] = UID
    rsp = requests.post(url, headers=headers, timeout=(10, 10), verify=False)
    logging.info(rsp.json())

    # 变量设备列表，筛选指定设备类型的设备名称并返回，例如，默认类型是V8P
    devices_list = list(rsp.json()["data"]["list"])

    # 设备列表非空，返回设备列表，设备列表为空，返回None
    if devices_list:
        return devices_list
    else:
        return None


def get_user_type(model='V8P', gz_host=gzHostCnTmp):
    """
    获取设备列表中指定的model的设备，并将其返回
    :param gz_host: 域名，默认为中国区
    :param model: 默认是V8P
    :return: 将设备的名称返回
    """
    pwd_md5 = _md5(pwd)
    _login(gz_host, _email=email, _region='CN', country_code='86', _password=pwd_md5, _type=1)
    url = 'https://' + gz_host + '/v1/cloud/getStatusList' + '?' + 'uuid=' + 'android_ui_auto' + '&' + 't=' + '005'
    headers = _headers()
    headers['Gz-Sid'] = SID
    headers['Gz-Uid'] = UID
    rsp = requests.post(url, headers=headers, timeout=(10, 10), verify=False)
    logging.info(rsp.json())

    # 变量设备列表，筛选指定设备类型的设备名称并返回，例如，默认类型是V8P
    user_type = rsp.json()["data"]["userType"]
    if user_type:
        return user_type


def aosu_admin_login(aosu_host='admin-cn.aosulife.com', pid='glazero', username='zhangjiamin', password='123'):
    pwd_md5 = _md5(password)
    headers = aosu_headers()
    url = 'https://' + aosu_host + '/admin/adminUser/login' + '?' + 'pid=' + pid + '&' + 'uuid=' + str(uuid.uuid1())
    data = 'pid=' + pid + '&' + 'username=' + username + '&' + 'password=' + pwd_md5 + '&' + 'uuid=' + str(uuid.uuid1())
    rsp = requests.post(url, headers=headers, data=data, timeout=(10, 10), verify=False)
    token = rsp.json()['data']['token']
    return token


def aosu_admin_get_dev_info(sn_sys='H1L2AH110000650', pid='glazero', gz_username='zhangjiamin',
                            aosu_host='admin-cn.aosulife.com'):
    # 获取token
    gz_sid = aosu_admin_login(aosu_host='admin-cn.aosulife.com', pid='glazero', username='zhangjiamin', password='123')
    headers = aosu_headers()
    url = 'https://' + aosu_host + '/admin/dev/getInfoList' + '?' + 'pid=' + pid + '&' + 'uuid=' + str(uuid.uuid1())
    data = 'pid=' + pid + '&' + 'gz_sid=' + gz_sid + '&' + 'gz_username=' + gz_username + '&' + 'snSys=' + sn_sys + \
           '&sn=&tuyaUuid=&devType=&' + 'uuid=' + str(uuid.uuid1())
    rsp = requests.post(url, headers=headers, data=data, timeout=(10, 10), verify=False)
    return rsp
    # print(rsp.json())
    # print("aosu状态为：", rsp.json()['data']['list'][0]['online'])
    # print("tuya状态为：", rsp.json()['data'x]['list'][0]['tuyayOnline'])


def get_dsc(device="SamsungA51"):
    path = os.path.dirname(os.path.dirname(os.path.realpath(__file__)))

    # windows系统
    if os.name == 'nt':
        yaml_url = os.path.join(path, 'gz_ui_auto\\devices.yaml')
    # macOS或者Linux
    elif os.name == 'posix':
        yaml_url = os.path.join(path, 'glazero_ui_auto/devices.yaml')
    else:
        raise RuntimeError("Unsupported operating system!")

    print("yaml配置文件地址：%s" % yaml_url)

    f = open(yaml_url, 'r', encoding='utf-8')
    file = f.read()
    f.close()
    data = yaml.load(file, Loader=yaml.FullLoader)
    for content in data:
        if device in content["desc"]:
            return content


def get_app_log(log_type, log_date, current_time, pull_local_path, numbers_of_lines=1000):
    """
    获取app日志或者涂鸦日志
    :param current_time: 取日志时的时间，格式：年月日-时分秒，用于命名文件，不同的附件要对应不同的文件
    :param log_type: app 或者 ty
    :param log_date: 日志文件中的日期例如，20230510，用于选择对应的日志文件
    :param numbers_of_lines：返回日志的行数，默认是最新的1000行
    :结果: 将重定向的日志文件pull到本地，作为allure的attachment
    """
    if log_type == 'app':
        file_name = 'aosu_app_android_' + str(log_date) + '.log'
    elif log_type == 'ty':
        file_name = 'aosu_app_android_ty_' + str(log_date) + '.log'

    # 获取device id
    cmd = 'adb devices'
    with os.popen(cmd, 'r') as f_log:
        devs_id = f_log.readlines()
        dev_id = re.findall(r'^\w*\b', devs_id[1])[0]

    # 进入adb shell后进入日志目录，获取对应日期和对应日志类型的的日志
    cmd = 'adb -s %s shell "cd /sdcard/Android/data/com.glazero.android/files/log && ls && cat %s | tail -n %d > ' \
          '%s_log_%s.log && ls"' % (dev_id, file_name, numbers_of_lines, log_type, current_time)
    os.system(cmd)

    # 将到出的日志pull到本地
    cmd = 'adb pull /sdcard/Android/data/com.glazero.android/files/log/%s_log_%s.log %s' % (
        log_type, current_time, pull_local_path)
    os.system(cmd)


def time_difference_medium(new_time, old_time):
    # 基于 datatime库 的时间戳
    # 计算出的时间时间戳 差值是 时分秒毫秒 格式
    # new_time 为字符串类型，old_time为 时间戳
    # 使用strptime解析字符串为datetime对象
    success_specified_time = datetime.strptime(new_time, "%Y-%m-%d %H:%M:%S.%f")
    print("指定的时间:", success_specified_time)
    # 计算从点击到获取字段成功时的时间差
    success_time_difference = success_specified_time - old_time
    print(success_time_difference)
    # 将时间差格式化为时分秒毫秒形式
    days, seconds = success_time_difference.days, success_time_difference.seconds
    hours = days * 24 + seconds // 3600
    minutes = (seconds % 3600) // 60
    seconds = seconds % 60
    milliseconds = success_time_difference.microseconds // 1000

    formatted_time = "{:02}:{:02}:{:02}.{:03}".format(hours, minutes, seconds, milliseconds)

    print("时间差（时分秒毫秒）:", formatted_time)

    return success_time_difference


def result_save_excel(file_path, Sheet, content_to_add, column_to_add):
    # 打开Excel文件
    file_path = file_path
    workbook = openpyxl.load_workbook(file_path)
    # 选择要操作的工作表（目前是Sheet1,默认Sheet1,注意大小写,如果表格中工作表名称是 Sheet1 ,写成小写的 sheet 代码会执行失败）
    sheet = workbook[Sheet]
    # 要添加的内容
    content_to_add = content_to_add
    # 获取列的最大行号
    max_row = sheet.max_row
    # 找到列中的最后一个非空单元格
    last_row = None
    # 要写入的列, ***** 注意:写入列column_to_add的类型为列表list !!!!!
    column_to_add = column_to_add
    # 如果所在列 行内,内容不为空,继续执行,直到行内容为空,跳出循环
    for row in range(max_row, 0, -1):
        # 获取 想要添加的列中 行数内的内容.
        cell_value = sheet[f'{column_to_add}{row}'].value
        # 如果所在列 行内,内容不为空, 将当前行数赋值给 loast_row 继续执行,直到行内容为空,即当前最后一行有数据的行数。跳出循环.
        if cell_value is not None:
            last_row = row
            break
    if last_row is not None:
        # 在列的最后一个非空单元格的下一行追加内容
        sheet[f'{column_to_add}{last_row + 1}'] = content_to_add

    # 保存修改后的Excel文件
    workbook.save(file_path)
    # 关闭 Excel 文件
    workbook.close()


def result_save_excel_column_list(file_path, Sheet, content_to_add, column_list_to_add):
    # 打开Excel文件
    file_path = file_path
    workbook = openpyxl.load_workbook(file_path)
    # 选择要操作的工作表（目前是Sheet1,默认Sheet1,注意大小写,如果表格中工作表名称是 Sheet1 ,写成小写的 sheet 代码会执行失败）
    sheet = workbook[Sheet]
    # 要添加的内容
    content_to_add = content_to_add
    # 获取列的最大行号
    max_row = sheet.max_row
    # 找到列中的最后一个非空单元格
    last_row = None
    # 要写入的列, ***** 注意:写入列column_to_add的类型为列表list !!!!!
    column_to_add = column_list_to_add
    # 如果所在列 行内,内容不为空,继续执行,直到行内容为空,跳出循环
    for item in column_list_to_add:
        column = item
        for row in range(max_row, 0, -1):
            # 获取 想要添加的列中 行数内的内容.
            cell_value = sheet[f'{column}{row}'].value
            # 如果所在列 行内,内容不为空, 将当前行数赋值给 loast_row 继续执行,直到行内容为空,即当前最后一行有数据的行数。跳出循环.
            if cell_value is not None:
                last_row = row
                break
        if last_row is not None:
            # 在列的最后一个非空单元格的下一行追加内容
            sheet[f'{column}{last_row + 1}'] = content_to_add
    # 保存修改后的Excel文件
    workbook.save(file_path)
    # 关闭 Excel 文件
    workbook.close()


def result_save_excel_full_list(file_path, Sheet, content_full_list_to_add, column_full_list_to_add):
    # 打开Excel文件
    file_path = file_path
    workbook = openpyxl.load_workbook(file_path)
    # 选择要操作的工作表（目前是Sheet1,默认Sheet1,注意大小写,如果表格中工作表名称是 Sheet1 ,写成小写的 sheet 代码会执行失败）
    sheet = workbook[Sheet]
    # 获取列的最大行号
    max_row = sheet.max_row
    # 找到列中的最后一个非空单元格
    last_row = None
    # 要写入的列, ***** 注意:写入列column_to_add的类型为列表list !!!!!
    column_full_list_to_add = column_full_list_to_add
    # 要添加的内容
    content_full_list_to_add = content_full_list_to_add
    # 如果所在列 行内,内容不为空,继续执行,直到行内容为空,跳出循环
    i = 0
    for item in column_full_list_to_add:
        column = item
        for row in range(max_row, 0, -1):
            # 获取 想要添加的列中 行数内的内容.
            cell_value = sheet[f'{column}{row}'].value
            # 如果所在列 行内,内容不为空, 将当前行数赋值给 loast_row 继续执行,直到行内容为空,即当前最后一行有数据的行数。跳出循环.
            if cell_value is not None:
                last_row = row
                break
        if last_row is not None:
            # 在列的最后一个非空单元格的下一行追加内容
            content = content_full_list_to_add[i]
            sheet[f'{column}{last_row + 1}'] = content
        i = i + 1
    # 保存修改后的Excel文件
    workbook.save(file_path)

    # 关闭 Excel 文件
    workbook.close()


def select_log_keyword_state_advanced(log_keyword_name, log_tag, log_keyword, log_keyword_contect, log_keyword_line,
                                      dev_id=initPhone.get_dev_id()):
    # adb shell logcat - v time - s TAG名 （-s表示只抓取这个TAG的日志）
    # dp = '\'\{\\"149\\":false\}\''
    cmd = 'adb -s %s shell logcat -v time -s %s "\| grep -e %s"' % (dev_id, log_tag, log_keyword)
    print(cmd)

    p_obj = subprocess.Popen(cmd, shell=False, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
    # 等待日志获取完成
    time.sleep(4)
    # 获取完成后药结束子进程
    p_obj.terminate()
    p_obj.kill()
    # 获取结束后读取内容
    lines = p_obj.stdout.readlines()
    print(lines)
    # 如果获取到的内容为空,执行返回值状态为空,否则,根据返回内容,返回成功或失败
    if len(lines) == 0:
        # 拿到的字段为空
        print(' %s is null' % log_keyword_name)
        # select_log_state_state = ' %s is null' % log_keyword_name  # 如果最后一行中,没有 要搜索的关键字  返回 关键词为空,供后面断言失败使用.
        select_log_state_state = []  # 如果最后一行中,没有 要搜索的关键字  返回 关键词为空,供后面断言失败使用.
        success_datatime = None
        '''
        # 将 从点击屏幕返回 到 设备休眠成功 的次数,存储在 ./data.xlsx 路径文件中的 'AO' 列最后一行
        colume_to_add_null = ['AO']
        result_save_excel_column_list('./data.xlsx', 'Sheet1', 0, colume_to_add_null)
        # 将 从点击屏幕返回 到 设备休眠成功 的时间,存储在 ./data.xlsx 路径文件中的 G 列最后一行
        colume_to_add_null = ['G']
        result_save_excel_column_list('./data.xlsx', 'Sheet1', '休眠状态为空', colume_to_add_null)
        '''
        # 获取完成后清除logcat
        r_obj = os.popen("adb logcat -c")
        r_obj.close()
    else:
        # 将bytes转换成字符串
        last_line = lines[log_keyword_line].decode('utf-8')
        print("搜索的日志关键字所在行的内容：", last_line)

        # 取日志中 设备唤醒成功这行的日期,即这行的 第 1 个字符串
        get_success_date = last_line.split(" ")[0]
        # 取日志中 设备唤醒成功这行的时间,即这行的 第 2 个字符串,时间中有毫秒
        get_success_time = last_line.split(" ")[1]

        # 获取当前年份
        # 获取当前时间 取年月日
        current_year = datetime.now().strftime("%Y")

        # 将从日志中取的时间 拼接成字符串.格式为 年月日 时分秒.
        success_datatime = current_year + '-' + get_success_date + ' ' + get_success_time
        print("日志中的时间：", success_datatime)

        '''
        # 计算  从点击屏幕返回 到 设备休眠成功 的时间差值
        time_difference_success = time_difference(last_line, click_time)
        '''
        results = last_line.split()
        print(results)
        # 查找 ensureDeviceWake wakeSuccess 字段
        for word in results:
            if '%s' % log_keyword_contect in word:
                # 如果查找的字段 查找成功 返回值为成功.
                select_log_state_state = word
                print(select_log_state_state)
                '''
                # 将 从点击屏幕返回 到 设备休眠成功 的时间,存储在 ./data.xlsx 路径文件中的 'AO' 列最后一行
                colume_to_add_success = ['AO']
                result_save_excel_column_list('./data.xlsx', 'Sheet1', 1, colume_to_add_success)
                # 将 从点击屏幕返回 到 设备休眠成功 的时间,存储在 ./data.xlsx 路径文件中的 G 列最后一行
                colume_to_add_success = ['G']
                result_save_excel_column_list('./data.xlsx', 'Sheet1', time_difference_success, colume_to_add_success)
                '''
                break
        else:
            select_log_state_state = ' %s is failes' % log_keyword_name  # 如果最后一行中,没有 '\"149\":false' 即 设备依然在唤醒状态 .
            '''
            # 将 从点击屏幕返回 到 设备休眠成功 的时间,存储在 ./data.xlsx 路径文件中的 'AO' 列最后一行
            colume_to_add_failed = ['AO']
            result_save_excel_column_list('./data.xlsx', 'Sheet1', 0, colume_to_add_failed)
            # 将 从点击屏幕返回 到 设备休眠成功 的时间,存储在 ./data.xlsx 路径文件中的 G 列最后一行
            colume_to_add_failed = ['G']
            result_save_excel_column_list('./data.xlsx', 'Sheet1', '休眠失败', colume_to_add_failed)
            '''
            # 获取完成后清除logcat
            r_obj = os.popen("adb logcat -c")
            r_obj.close()

    # # 获取完成后清除logcat
    # r_obj = os.popen("adb logcat -c")
    # r_obj.close()

    return select_log_state_state, success_datatime


def select_log_keyword_state_medium(log_keyword_name, log_tag, log_keyword, log_keyword_contect, log_keyword_line,
                                    dev_id=initPhone.get_dev_id()):
    # adb shell logcat - v time - s TAG名 （-s表示只抓取这个TAG的日志）
    # dp = '\'\{\\"149\\":false\}\''
    cmd = 'adb -s %s shell logcat -v time -s %s "\| grep -e %s"' % (dev_id, log_tag, log_keyword)
    print(cmd)

    p_obj = subprocess.Popen(cmd, shell=False, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
    # 等待日志获取完成
    time.sleep(4)
    # 获取完成后药结束子进程
    p_obj.terminate()
    p_obj.kill()
    # 获取结束后读取内容
    lines = p_obj.stdout.readlines()
    print(lines)
    # 如果获取到的内容为空,执行返回值状态为空,否则,根据返回内容,返回成功或失败
    if len(lines) == 0:
        select_log_state_state = ' %s is null' % log_keyword_name  # 如果最后一行中,没有 要搜索的关键字  返回 关键词为空,供后面断言失败使用.
        '''
        # 将 从点击屏幕返回 到 设备休眠成功 的次数,存储在 ./data.xlsx 路径文件中的 'AO' 列最后一行
        colume_to_add_null = ['AO']
        result_save_excel_column_list('./data.xlsx', 'Sheet1', 0, colume_to_add_null)
        # 将 从点击屏幕返回 到 设备休眠成功 的时间,存储在 ./data.xlsx 路径文件中的 G 列最后一行
        colume_to_add_null = ['G']
        result_save_excel_column_list('./data.xlsx', 'Sheet1', '休眠状态为空', colume_to_add_null)
        '''
        # 获取完成后清除logcat
        r_obj = os.popen("adb logcat -c")
        r_obj.close()
    else:
        # 将bytes转换成字符串
        last_line = lines[log_keyword_line].decode('utf-8')
        print("搜索的日志关键字所在行的内容：", last_line)

        '''
        # 计算  从点击屏幕返回 到 设备休眠成功 的时间差值
        time_difference_success = time_difference(last_line, click_time)
        '''
        results = last_line.split()
        print(results)
        # 查找 ensureDeviceWake wakeSuccess 字段
        for word in results:
            if '%s' % log_keyword_contect in word:
                select_log_state_state = word
                '''
                # 将 从点击屏幕返回 到 设备休眠成功 的时间,存储在 ./data.xlsx 路径文件中的 'AO' 列最后一行
                colume_to_add_success = ['AO']
                result_save_excel_column_list('./data.xlsx', 'Sheet1', 1, colume_to_add_success)
                # 将 从点击屏幕返回 到 设备休眠成功 的时间,存储在 ./data.xlsx 路径文件中的 G 列最后一行
                colume_to_add_success = ['G']
                result_save_excel_column_list('./data.xlsx', 'Sheet1', time_difference_success, colume_to_add_success)
                '''
                break
        else:
            select_log_state_state = ' %s is failes' % log_keyword_name  # 如果最后一行中,没有 '\"149\":false' 即 设备依然在唤醒状态 .
            '''
            # 将 从点击屏幕返回 到 设备休眠成功 的时间,存储在 ./data.xlsx 路径文件中的 'AO' 列最后一行
            colume_to_add_failed = ['AO']
            result_save_excel_column_list('./data.xlsx', 'Sheet1', 0, colume_to_add_failed)
            # 将 从点击屏幕返回 到 设备休眠成功 的时间,存储在 ./data.xlsx 路径文件中的 G 列最后一行
            colume_to_add_failed = ['G']
            result_save_excel_column_list('./data.xlsx', 'Sheet1', '休眠失败', colume_to_add_failed)
            '''
            # 获取完成后清除logcat
            r_obj = os.popen("adb logcat -c")
            r_obj.close()

    # # 获取完成后清除logcat
    # r_obj = os.popen("adb logcat -c")
    # r_obj.close()

    return select_log_state_state


def select_log_keyword_state_simple(log_keyword_name, log_tag, log_keyword_line, dev_id=initPhone.get_dev_id()):
    # adb shell logcat - v time - s TAG名 （-s表示只抓取这个TAG的日志）
    # dp = '\'\{\\"149\\":false\}\''
    cmd = 'adb -s %s shell logcat -v time -s %s' % (dev_id, log_tag)
    print(cmd)

    p_obj = subprocess.Popen(cmd, shell=False, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
    # 等待日志获取完成
    time.sleep(4)
    # 获取完成后药结束子进程
    p_obj.terminate()
    p_obj.kill()
    # 获取结束后读取内容
    lines = p_obj.stdout.readlines()
    print(lines)
    # 如果获取到的内容为空,执行返回值状态为空,否则,根据返回内容,返回成功或失败
    if len(lines) == 0:
        select_log_state = ' %s is null' % log_keyword_name  # 如果最后一行中,没有 要搜索的关键字  返回 关键词为空,供后面断言失败使用.
        '''
        # 将 从点击屏幕返回 到 设备休眠成功 的次数,存储在 ./data.xlsx 路径文件中的 'AO' 列最后一行
        colume_to_add_null = ['AO']
        result_save_excel_column_list('./data.xlsx', 'Sheet1', 0, colume_to_add_null)
        # 将 从点击屏幕返回 到 设备休眠成功 的时间,存储在 ./data.xlsx 路径文件中的 G 列最后一行
        colume_to_add_null = ['G']
        result_save_excel_column_list('./data.xlsx', 'Sheet1', '休眠状态为空', colume_to_add_null)
        '''
        # 获取完成后清除logcat
        r_obj = os.popen("adb logcat -c")
        r_obj.close()
    else:
        # 将bytes转换成字符串
        last_line = lines[{log_keyword_line}].decode('utf-8')
        print("搜索的日志关键字所在行的内容：", last_line)
        '''
        # 计算  从点击屏幕返回 到 设备休眠成功 的时间差值
        time_difference_success = time_difference(last_line, click_time)
        '''
        results = last_line.split()
        print(results)
        results = last_line.split(" ")[-1]
        # 查找 ensureDeviceWake wakeSuccess 字段
        for word in results:
            if '%s' % log_keyword_name in word:
                select_log_state = word
                '''
                # 将 从点击屏幕返回 到 设备休眠成功 的时间,存储在 ./data.xlsx 路径文件中的 'AO' 列最后一行
                colume_to_add_success = ['AO']
                result_save_excel_column_list('./data.xlsx', 'Sheet1', 1, colume_to_add_success)
                # 将 从点击屏幕返回 到 设备休眠成功 的时间,存储在 ./data.xlsx 路径文件中的 G 列最后一行
                colume_to_add_success = ['G']
                result_save_excel_column_list('./data.xlsx', 'Sheet1', time_difference_success, colume_to_add_success)
                '''

            if 'PlaybackActivity' in word:
                break

        else:
            select_log_state = ' %s is failes' % log_keyword_name  # 如果最后一行中,没有 '\"149\":false' 即 设备依然在唤醒状态 .
            '''
            # 将 从点击屏幕返回 到 设备休眠成功 的时间,存储在 ./data.xlsx 路径文件中的 'AO' 列最后一行
            colume_to_add_failed = ['AO']
            result_save_excel_column_list('./data.xlsx', 'Sheet1', 0, colume_to_add_failed)
            # 将 从点击屏幕返回 到 设备休眠成功 的时间,存储在 ./data.xlsx 路径文件中的 G 列最后一行
            colume_to_add_failed = ['G']
            result_save_excel_column_list('./data.xlsx', 'Sheet1', '休眠失败', colume_to_add_failed)
            '''
            # 获取完成后清除logcat
            r_obj = os.popen("adb logcat -c")
            r_obj.close()

    # 获取完成后清除logcat
    r_obj = os.popen("adb logcat -c")
    r_obj.close()

    return select_log_state


def time_format_conversion_datetime_format(new_time):
    # 获取当前时间 取年月日
    current_today = date.today().strftime("%Y-%m-%d ")

    # 将从日志中取的时间 拼接成字符串.格式为 年月日 时分秒.
    success_datatime = current_today + new_time
    print("日志中的时间：", success_datatime)

    # 使用strptime解析字符串为datetime对象
    success_specified_time = datetime.strptime(success_datatime, "%Y-%m-%d %H:%M:%S")
    print("指定的时间:", success_specified_time)
    print(type(success_specified_time))

    # 将 年月日时分秒 格式 转化成  时间戳
    success_time_timestamp_int = int(success_specified_time.timestamp())
    print(success_time_timestamp_int)
    print(type(success_time_timestamp_int))

    success_time_timestamp = datetime.fromtimestamp(success_time_timestamp_int)
    print(success_time_timestamp)
    print(type(success_time_timestamp))

    return success_time_timestamp


def time_difference_simple(new_time, old_time):
    # 获取当前时间 取年月日
    current_today = date.today().strftime("%Y-%m-%d ")

    # 将从日志中取的时间 拼接成字符串.格式为 年月日 时分秒.
    success_datatime = current_today + new_time
    print("日志中的时间：", success_datatime)

    # 使用strptime解析字符串为datetime对象
    success_specified_time = datetime.strptime(success_datatime, "%Y-%m-%d %H:%M:%S")
    print("指定的时间:", success_specified_time)

    # 计算从点击到获取字段成功时的时间差 (atetime.timedelta 类型 时间类型)
    success_time_difference = success_specified_time - old_time
    print(success_time_difference)

    return success_time_difference


def string_change_datetime(time_str):
    # 将 字符串类型 时间,转化成 datetime类型,datetime 也可用于datetime类型的 时间加减.
    # 使用strptime解析字符串为datetime对象
    success_specified_time = datetime.strptime(time_str, "%Y-%m-%d %H:%M:%S.%f")
    print("字符串转成datetime类型的时间:", success_specified_time)
    print(type(success_specified_time))

    # # 将 年月日时分秒 格式 转化成  时间戳
    # success_time_timestamp_int = int(success_specified_time.timestamp()) * 1000
    # print(success_time_timestamp_int)
    # print(type(success_time_timestamp_int))
    #
    # success_time_timestamp = datetime.fromtimestamp(success_time_timestamp_int)
    # print(success_time_timestamp)
    # print(type(success_time_timestamp))

    return success_specified_time


def datetime_timedelta_change_string(datetime_timedelta):
    # 将时间差格式化为时分秒毫秒形式
    days, seconds = datetime_timedelta.days, datetime_timedelta.seconds
    hours = days * 24 + seconds // 3600
    minutes = (seconds % 3600) // 60
    seconds = seconds % 60
    milliseconds = datetime_timedelta.microseconds // 1000

    formatted_time = "{:02}:{:02}:{:02}.{:03}".format(hours, minutes, seconds, milliseconds)

    print("时间差（时分秒毫秒）:", formatted_time)

    return formatted_time


def log_upload_html(save_log_path):
    # 在手机中查找对应的日志文件
    current_date = time.strftime("%Y%m%d", time.localtime())

    # 获取当前时间，用于区分不同的日志文件作为附件。
    current_time = time.strftime("%Y-%m-%d_%H-%M-%S", time.localtime())

    # 获取app日志
    get_app_log('app', current_date, current_time, './report/%s/log_attch' % save_log_path, 1000)

    # 将日志添加到报告中
    allure.attach.file("./report/%s/log_attch/app_log_%s.log" % (save_log_path, current_time), name="app log",
                       attachment_type=allure.attachment_type.TEXT)
    time.sleep(2)

    # 获取ty日志
    get_app_log('ty', current_date, current_time, './report/%s/log_attch' % save_log_path, 4000)
    # 将日志添加到报告中
    allure.attach.file("./report/%s/log_attch/ty_log_%s.log" % (save_log_path, current_time), name="ty log",
                       attachment_type=allure.attachment_type.TEXT)
    time.sleep(2)


def picture_contrast_black_and_white(image_path_name, image_black_and_white_path_name):
    # 打开图像文件
    image = Image.open(image_path_name)

    # 将图片转换为灰度图像
    gray_img = image.convert("L")

    # 调整图像对比度
    partial_screenshot_contrast_img = ImageOps.autocontrast(gray_img, cutoff=0)

    # 保存黑白 截图照片
    partial_screenshot_contrast_img.save(image_black_and_white_path_name)


def get_image_text_simple(image_path_name, keyword_line, x0, y0, x1, y1, image_partial_path_name):
    # image_path_name 要提取 图片的 相对位置或绝对位置.
    # 相对位置,例如:'./image/05.jpg'或绝对位置,例如:'C:/Users/L/PycharmProjects/gz_ui_auto/image/05.jpg'
    # 打开图像文件
    image = Image.open(image_path_name)

    # 获取屏幕分辨率
    width, height = image.size
    print(width, height)

    # 截取部分屏幕
    # left = 600  # int(width * 0.5)
    # top = 200  # int(height * 0.5)
    # right = 1080  # int(width * 0.8)
    # bottom = 250  # int(height * 0.8)
    # 截取部分图片,范围矩形:左上和游侠坐标
    partial_screenshot = image.crop((x0, y0, x1, y1))
    # 保存黑白 截图照片
    partial_screenshot.save(image_partial_path_name)
    # Rect_coord = '(x0 ,y0 ,x1 , y1)'

    # 使用python-tesseract进行文本识别
    # 只匹配字符串就不要加 lang= ,会报错.  如果匹配内容包含中文,加上 lang= chi_sim 代表的简体中文,如果匹配结果不含中文,可以不加
    text = pytesseract.image_to_string(partial_screenshot)
    print(text)

    # 去除识别结果中的非法字符
    cleaned_text = re.sub(r'[^a-zA-Z0-9:\-\s]+?', '', text)

    # 将识别结果按行分割成列表
    lines = text.split('\n')  # eg:['16:04 Ml Q% - On', '< panda O', '', '2024-01-03 Wed 16:04:30', '76KB/S', 'E', '']
    print(lines)

    # 去除空行
    lines = [line.strip() for line in lines if line.strip()]
    print(lines)  # ['16:04 Ml Q% - On', '< panda O', '2024-01-03 Wed 16:04:30', '76KB/S','76KB/S', 'E',]

    # 获取 改行的第几个字段
    try:
        line = lines[keyword_line]
        print(line)

        # 取出 图片中识别到的文字: 时间戳
        text = line.split(",")[0]
        print(text)
    except:
        text = ""
        print(text)

    # 返回图片识别到的内容
    return text


def kill_app_process(devices_name, app_id):
    # adb shell logcat - v time - s TAG名 （-s表示只抓取这个TAG的日志）
    # dp = '\'\{\\"149\\":false\}\''
    cmd = 'adb -s %s shell am force-stop %s' % (devices_name, app_id)
    print(cmd)

    os.system(cmd)


def get_toast_text(driver, toast_text):
    # 定位（捕获）toast元素
    # 定位器
    locator = ("xpath", "//*[contains(@text,'邀请已发送')]")
    # 定位
    toast = WebDriverWait(driver, 10, 0.01).until(EC.presence_of_element_located(locator))

    # 输出toast信息
    print(toast.text)

    return toast.text
