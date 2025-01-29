import os
import re
import time
import subprocess
import time
import gz_public
import xlwt
import xlrd
from xlutils.copy import copy
import openpyxl
from datetime import datetime


def get_dev_id():
    cmd = 'adb devices'
    with os.popen(cmd, 'r') as f_obj:
        devs_id = f_obj.readlines()
        dev_id = re.findall(r'^\w*\b', devs_id[1])[0]
    return dev_id


def get_dev_play_state(dev_id=get_dev_id()):
    cmd = 'adb -s %s shell logcat -v time -s LivePlayer "\| grep -e state"' % dev_id
    print(cmd)

    p_obj = subprocess.Popen(cmd, shell=False, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
    # 等待日志获取完成
    time.sleep(4)
    # 获取完成后药结束子进程
    p_obj.terminate()
    p_obj.kill()
    # 获取结束后读取内容
    lines = p_obj.stdout.readlines()
    # 将bytes转换成字符串
    last_line = lines[-1].decode('utf-8')
    print("最后一行：", last_line)
    results = last_line.split()

    # 查找playState字段
    for word in results:
        if 'state' in word:
            play_state = word
            break

    # 获取完成后清除logcat
    r_obj = os.popen("adb logcat -c")
    r_obj.close()

    return play_state


def get_start_create_LivePlay_fragment_result(dev_id, click_time):
    # 获取 APP 从首页进入开流页面时  的日志
    cmd = 'adb -s %s shell logcat -v time -s BaseFragment "\| grep -e LivePlaySingleFragment:onCreate"' % dev_id  # C9S  LivePlayFragment:onCreate改成了:LivePlaySingleFragment:onCreate
    print(cmd)

    p_obj = subprocess.Popen(cmd, shell=False, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
    # 等待日志获取完成
    time.sleep(4)
    # 获取完成后药结束子进程
    p_obj.terminate()
    p_obj.kill()
    # 获取结束后读取内容
    lines = p_obj.stdout.readlines()
    print(lines)
    # 如果获取到的内容为空执行返回值状态为空,否则,根据返回内容,返回成功或失败
    if len(lines) == 0:
        create_LivePlay_fragment_state = 'create_fragment:null'
        print('开流播放片段在创建状态为空')
        # 将从点击屏幕 到 进入开流页面 的次数,存储在 ./data.xlsx 路径文件中的 'AJ', 'AK', 'AL', 'AM', 'AN', 'AO' 列最后一行
        colume_list_to_add_null = ['AJ', 'AK', 'AL', 'AM', 'AN', 'AO']
        result_save_excel_column_list('./data.xlsx', 'Sheet1', 0, colume_list_to_add_null)
        # 将从点击屏幕 到 进入开流页面 的时间,存储在 ./data.xlsx 路径文件中的 'B', 'C', 'D', 'E', 'F', 'G' 列最后一行
        colume_to_add_null = ['B', 'C', 'D', 'E', 'F', 'G']
        content_to_add_null = ['进入开流页面为空', '开始唤醒为空', '唤醒为空', 'p2p连接为空', 'Preview 为空',
                               '休眠为空']
        result_save_excel_full_list('./data.xlsx', 'Sheet1', content_to_add_null, colume_to_add_null)

        # 获取完成后清除logcat
        r_obj = os.popen("adb logcat -c")
        r_obj.close()
    else:
        # 将bytes转换成字符串
        last_line = lines[-2].decode('utf-8')  # -1 改成 -2
        print("最后一行：", last_line)
        # 计算 从点击屏幕 到 进入开流页面 的时间差值
        time_difference_success = time_difference(last_line, click_time)
        # 将 last_line 的内容赋值给 results
        results = last_line.split()
        print(results)
        # 查找 onCreate 字段
        for word in results:
            if 'LivePlaySingleFragment:onCreate' in word:
                create_LivePlay_fragment_state = word
                print('开始创建开流现场播放片段(进入开流页面):', create_LivePlay_fragment_state)

                # 将从点击屏幕开流 到 进入开流页面 的次数,存储在 ./data.xlsx 路径文件中的 AJ列最后一行
                colume_to_add_success = ['AJ']
                result_save_excel_column_list('./data.xlsx', 'Sheet1', 1, colume_to_add_success)
                # 将从点击屏幕开流 到 进入开流页面 的时间,存储在 ./data.xlsx 路径文件中的 AB 列最后一行
                colume_to_add_success = ['B']
                result_save_excel_column_list('./data.xlsx', 'Sheet1', time_difference_success, colume_to_add_success)

                break
        else:
            create_LivePlay_fragment_state = 'create_fragment:Failed'  # 如果最后一行中,没有wakeSuccess即 设备唤醒 失败, dev_wake_state 赋值为失败状态.
            print('设备未收到唤醒指令:', create_LivePlay_fragment_state)

            # 将从点击屏幕开流 到 进入开流页面 的次数,存储在 ./data.xlsx 路径文件中的 'AJ', 'AK', 'AL', 'AM', 'AN', 'AO' 列最后一行
            colume_to_add_failed = ['AJ', 'AK', 'AL', 'AM', 'AN', 'AO']
            result_save_excel_column_list('./data.xlsx', 'Sheet1', 0, colume_to_add_failed)
            # 将从点击屏幕开流 到 进入开流页面 的时间,存储在 ./data.xlsx 路径文件中的 'B', 'C', 'D', 'E', 'F', 'G' 列最后一行
            colume_to_add_failed = ['B', 'C', 'D', 'E', 'F', 'G']
            content_to_add_failed = ['进入开流页面失败', '开始唤醒失败', '唤醒失败', 'p2p连接失败', 'Preview 失败',
                                     '休眠失败']
            result_save_excel_full_list('./data.xlsx', 'Sheet1', content_to_add_failed, colume_to_add_failed)

            # 获取完成后清除logcat
            r_obj = os.popen("adb logcat -c")
            r_obj.close()

    return create_LivePlay_fragment_state


def get_dev_start_wake_state_result(dev_id, click_time):
    # 获取 设备收到 开始唤醒时 的日志
    cmd = 'adb -s %s shell logcat -v time -s TyCameraCenter "\| grep -e wakeStart"' % dev_id
    print(cmd)

    p_obj = subprocess.Popen(cmd, shell=False, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
    # 等待日志获取完成
    time.sleep(4)
    # 获取完成后药结束子进程
    p_obj.terminate()
    p_obj.kill()
    # 获取结束后读取内容
    lines = p_obj.stdout.readlines()
    print(lines)
    # 如果获取到的内容为空执行返回值状态为空,否则,根据返回内容,返回成功或失败
    if len(lines) == 0:
        start_wake_state = 'startwakeState:null'
        print('唤醒状态为空')

        # 将从 点击屏幕开流 到 设备收到 开始唤醒时 的次数,存储在 ./data.xlsx 路径文件中的  'AK', 'AL', 'AM', 'AN', 'AO' 列最后一行
        colume_list_to_add_null = ['AK', 'AL', 'AM', 'AN', 'AO']
        result_save_excel_column_list('./data.xlsx', 'Sheet1', 0, colume_list_to_add_null)
        # 将从 点击屏幕开流 到 设备收到 开始唤醒时 的时间,存储在 ./data.xlsx 路径文件中的  'C', 'D', 'E', 'F', 'G' 列最后一行
        colume_to_add_null = ['C', 'D', 'E', 'F', 'G']
        content_to_add_null = ['开始唤醒为空', '唤醒为空', 'p2p连接为空', 'Preview 为空', '休眠为空']
        result_save_excel_full_list('./data.xlsx', 'Sheet1', content_to_add_null, colume_to_add_null)

        # 获取完成后清除logcat
        r_obj = os.popen("adb logcat -c")
        r_obj.close()
    else:
        # 将bytes转换成字符串
        last_line = lines[-1].decode('utf-8')  # 1 改成 -1
        print("最后一行：", last_line)
        # 计算 从 点击屏幕开流 到 设备收到 开始唤醒时 的时间差值
        time_difference_success = time_difference(last_line, click_time)
        # 将 last_line 的内容赋值给 results
        results = last_line.split()
        print(results)
        # 查找 ensureDeviceWake 字段
        for word in results:
            if 'wakeStart' in word:
                start_wake_state = word
                print('设备收到开始唤醒指令:', start_wake_state)

                # 将从 点击屏幕开流 到 设备收到 开始唤醒时 的次数,存储在 ./data.xlsx 路径文件中的 AK列最后一行
                colume_to_add_success = ['AK']
                result_save_excel_column_list('./data.xlsx', 'Sheet1', 1, colume_to_add_success)

                # 将从 点击屏幕开流 到 设备收到 开始唤醒时 的时间,存储在 ./data.xlsx 路径文件中的 AC 列最后一行
                colume_to_add_success = ['C']
                result_save_excel_column_list('./data.xlsx', 'Sheet1', time_difference_success, colume_to_add_success)

                break
        else:
            start_wake_state = 'startwakeFailed'  # 如果最后一行中,没有 wakeSuccess 即 设备唤醒 失败, dev_wake_state 赋值为失败状态.
            print('设备未收到唤醒指令:', start_wake_state)

            # 将从 点击屏幕开流 到 设备收到 开始唤醒时 的次数,存储在 ./data.xlsx 路径文件中的 'AK', 'AL', 'AM', 'AN', 'AO' 列最后一行
            colume_to_add_failed = ['AK', 'AL', 'AM', 'AN', 'AO']
            result_save_excel_column_list('./data.xlsx', 'Sheet1', 0, colume_to_add_failed)
            # 将从 点击屏幕开流 到 设备收到 开始唤醒时 的时间,存储在 ./data.xlsx 路径文件中的 'C', 'D', 'E', 'F', 'G' 列最后一行
            colume_to_add_failed = ['C', 'D', 'E', 'F', 'G']
            content_to_add_failed = ['开始唤醒失败', '唤醒失败', 'p2p连接失败', 'Preview 失败', '休眠失败']
            result_save_excel_full_list('./data.xlsx', 'Sheet1', content_to_add_failed, colume_to_add_failed)

            # 获取完成后清除logcat
            r_obj = os.popen("adb logcat -c")
            r_obj.close()

    return start_wake_state


def get_dev_wake_state_result(dev_id, click_time):
    cmd = 'adb -s %s shell logcat -v time -s TyCameraCenter "\| grep -e ensureDeviceWake wakeSuccess"' % dev_id
    print(cmd)

    p_obj = subprocess.Popen(cmd, shell=False, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
    # 等待日志获取完成
    time.sleep(4)
    # 获取完成后药结束子进程
    p_obj.terminate()
    p_obj.kill()
    # 获取结束后读取内容
    lines = p_obj.stdout.readlines()
    print(lines)
    # 如果获取到的内容为空执行返回值状态为空,否则,根据返回内容,返回成功或失败
    if len(lines) == 0:
        dev_wake_state = 'wakeState:null'
        print('唤醒状态为空')
        # 将从点击屏幕开流到设备唤醒成功的次数,存储在 ./data.xlsx 路径文件中的 'AL', 'AM', 'AN', 'AO' 列最后一行
        colume_list_to_add_null = ['AL', 'AM', 'AN', 'AO']
        result_save_excel_column_list('./data.xlsx', 'Sheet1', 0, colume_list_to_add_null)
        # 将从点击屏幕开流到设备唤醒成功的时间,存储在 ./data.xlsx 路径文件中的 B、C、D、E 列最后一行
        colume_to_add_null = ['D', 'E', 'F', 'G']
        content_to_add_null = ['唤醒为空', 'p2p连接为空', 'Preview 为空', '休眠为空']
        result_save_excel_full_list('./data.xlsx', 'Sheet1', content_to_add_null, colume_to_add_null)

        # 获取完成后清除logcat
        r_obj = os.popen("adb logcat -c")
        r_obj.close()
    else:
        # 将bytes转换成字符串
        last_line = lines[-1].decode('utf-8')
        print("最后一行：", last_line)
        # 计算 从点击屏幕到唤醒成功的时间差值
        time_difference_success = time_difference(last_line, click_time)
        # 将 last_line 的内容赋值给 results
        results = last_line.split()
        print(results)
        # 查找 wakeSuccess 字段
        for word in results:
            if 'wakeSuccess' in word:
                dev_wake_state = word
                print('唤醒成功:', dev_wake_state)
                # 将从点击屏幕开流到设备唤醒成功的时间,存储在 ./data.xlsx 路径文件中的 AL列最后一行
                colume_to_add_success = ['AL']
                result_save_excel_column_list('./data.xlsx', 'Sheet1', 1, colume_to_add_success)
                # 将从点击屏幕开流到设备唤醒成功的时间,存储在 ./data.xlsx 路径文件中的 AD 列最后一行
                colume_to_add_success = ['D']
                result_save_excel_column_list('./data.xlsx', 'Sheet1', time_difference_success, colume_to_add_success)
                break
        else:
            dev_wake_state = 'wakeFailed'  # 如果最后一行中,没有wakeSuccess即 设备唤醒 失败, dev_wake_state 赋值为失败状态.
            print('唤醒失败:', dev_wake_state)
            # 将从点击屏幕开流到设备唤醒成功的时间,存储在 ./data.xlsx 路径文件中的 'AL', 'AM', 'AN', 'AO' 列最后一行
            colume_to_add_failed = ['AL', 'AM', 'AN', 'AO']
            result_save_excel_column_list('./data.xlsx', 'Sheet1', 0, colume_to_add_failed)
            # 将从点击屏幕开流到设备唤醒成功的时间,存储在 ./data.xlsx 路径文件中的 B、C、D、E 列最后一行
            colume_to_add_failed = ['D', 'E', 'F', 'G']
            content_to_add_failed = ['唤醒失败', 'p2p连接失败', 'Preview 失败', '休眠失败']
            result_save_excel_full_list('./data.xlsx', 'Sheet1', content_to_add_failed, colume_to_add_failed)

            # 获取完成后清除logcat
            r_obj = os.popen("adb logcat -c")
            r_obj.close()

    return dev_wake_state


def time_difference(last_line, click_time):
    # 基于 datatime库 的时间戳
    # 计算出的时间时间戳 差值是 时分秒毫秒 格式
    # 取日志中 设备唤醒成功这行的日期,即这行的 第 1 个字符串
    get_success_date = last_line.split(" ")[0]
    # 取日志中 设备唤醒成功这行的时间,即这行的 第 2 个字符串,时间中有毫秒
    get_success_time = last_line.split(" ")[1]

    # 将从日志中取的时间 拼接成字符串.格式为 年月日 时分秒.
    success_datatime = '2024-' + get_success_date + ' ' + get_success_time
    print("日志中的时间：", success_datatime)

    # 使用strptime解析字符串为datetime对象
    success_specified_time = datetime.strptime(success_datatime, "%Y-%m-%d %H:%M:%S.%f")
    print("指定的时间:", success_specified_time)
    # 计算从点击到获取字段成功时的时间差
    success_time_difference = success_specified_time - click_time
    print(success_time_difference)
    # 将时间差格式化为时分秒毫秒形式
    days, seconds = success_time_difference.days, success_time_difference.seconds
    hours = days * 24 + seconds // 3600
    minutes = (seconds % 3600) // 60
    seconds = seconds % 60
    milliseconds = success_time_difference.microseconds // 1000

    formatted_time = "{:02}:{:02}:{:02}.{:03}".format(hours, minutes, seconds, milliseconds)

    print("时间差（时分秒毫秒）:", formatted_time)

    return success_time_difference


def result_save_excel(file_path, Sheet, content_to_add, column_to_add):
    # 打开Excel文件
    file_path = file_path
    workbook = openpyxl.load_workbook(file_path)
    # 选择要操作的工作表（目前是Sheet1,默认Sheet1,注意大小写,如果表格中工作表名称是 Sheet1 ,写成小写的 sheet 代码会执行失败）
    sheet = workbook[Sheet]
    # 要添加的内容
    content_to_add = content_to_add
    # 获取列的最大行号
    max_row = sheet.max_row
    # 找到列中的最后一个非空单元格
    last_row = None
    # 要写入的列, ***** 注意:写入列column_to_add的类型为列表list !!!!!
    column_to_add = column_to_add
    # 如果所在列 行内,内容不为空,继续执行,直到行内容为空,跳出循环
    for row in range(max_row, 0, -1):
        # 获取 想要添加的列中 行数内的内容.
        cell_value = sheet[f'{column_to_add}{row}'].value
        # 如果所在列 行内,内容不为空, 将当前行数赋值给 loast_row 继续执行,直到行内容为空,即当前最后一行有数据的行数。跳出循环.
        if cell_value is not None:
            last_row = row
            break
    if last_row is not None:
        # 在列的最后一个非空单元格的下一行追加内容
        sheet[f'{column_to_add}{last_row + 1}'] = content_to_add

    # 保存修改后的Excel文件
    workbook.save(file_path)
    # 关闭 Excel 文件
    workbook.close()


def result_save_excel_column_list(file_path, Sheet, content_to_add, column_list_to_add):
    # 打开Excel文件
    file_path = file_path
    workbook = openpyxl.load_workbook(file_path)
    # 选择要操作的工作表（目前是Sheet1,默认Sheet1,注意大小写,如果表格中工作表名称是 Sheet1 ,写成小写的 sheet 代码会执行失败）
    sheet = workbook[Sheet]
    # 要添加的内容
    content_to_add = content_to_add
    # 获取列的最大行号
    max_row = sheet.max_row
    # 找到列中的最后一个非空单元格
    last_row = None
    # 要写入的列, ***** 注意:写入列column_to_add的类型为列表list !!!!!
    column_to_add = column_list_to_add
    # 如果所在列 行内,内容不为空,继续执行,直到行内容为空,跳出循环
    for item in column_list_to_add:
        column = item
        for row in range(max_row, 0, -1):
            # 获取 想要添加的列中 行数内的内容.
            cell_value = sheet[f'{column}{row}'].value
            # 如果所在列 行内,内容不为空, 将当前行数赋值给 loast_row 继续执行,直到行内容为空,即当前最后一行有数据的行数。跳出循环.
            if cell_value is not None:
                last_row = row
                break
        if last_row is not None:
            # 在列的最后一个非空单元格的下一行追加内容
            sheet[f'{column}{last_row + 1}'] = content_to_add
    # 保存修改后的Excel文件
    workbook.save(file_path)
    # 关闭 Excel 文件
    workbook.close()


def result_save_excel_full_list(file_path, Sheet, content_full_list_to_add, column_full_list_to_add):
    # 打开Excel文件
    file_path = file_path
    workbook = openpyxl.load_workbook(file_path)
    # 选择要操作的工作表（目前是Sheet1,默认Sheet1,注意大小写,如果表格中工作表名称是 Sheet1 ,写成小写的 sheet 代码会执行失败）
    sheet = workbook[Sheet]
    # 获取列的最大行号
    max_row = sheet.max_row
    # 找到列中的最后一个非空单元格
    last_row = None
    # 要写入的列, ***** 注意:写入列column_to_add的类型为列表list !!!!!
    column_full_list_to_add = column_full_list_to_add
    # 要添加的内容
    content_full_list_to_add = content_full_list_to_add
    # 如果所在列 行内,内容不为空,继续执行,直到行内容为空,跳出循环
    i = 0
    for item in column_full_list_to_add:
        column = item
        for row in range(max_row, 0, -1):
            # 获取 想要添加的列中 行数内的内容.
            cell_value = sheet[f'{column}{row}'].value
            # 如果所在列 行内,内容不为空, 将当前行数赋值给 loast_row 继续执行,直到行内容为空,即当前最后一行有数据的行数。跳出循环.
            if cell_value is not None:
                last_row = row
                break
        if last_row is not None:
            # 在列的最后一个非空单元格的下一行追加内容
            content = content_full_list_to_add[i]
            sheet[f'{column}{last_row + 1}'] = content
        i = i + 1
    # 保存修改后的Excel文件
    workbook.save(file_path)

    # 关闭 Excel 文件
    workbook.close()


def get_dev_p2p_state_result(dev_id, click_time):
    cmd = 'adb -s %s shell logcat -v time -s TyCameraCenter "\| grep -e connectP2pEnd success"' % dev_id
    print(cmd)

    p_obj = subprocess.Popen(cmd, shell=False, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
    print(p_obj)
    # 等待日志获取完成
    time.sleep(4)
    # 获取完成后要结束子进程
    p_obj.terminate()
    p_obj.kill()
    # 获取结束后读取内容
    lines = p_obj.stdout.readlines()
    print(lines)
    # 如果获取到的内容为空,执行返回值状态为空,否则,根据返回内容,返回成功或失败
    if len(lines) == 0:
        dev_p2p_state = 'p2pState:null'
        # 将 从点击屏幕开流 到 p2p唤醒成功 的次数,存储在 ./data.xlsx 路径文件中的 'AM', 'AN', 'AO' 列最后一行
        colume_to_add_null = ['AM', 'AN', 'AO']
        result_save_excel_column_list('./data.xlsx', 'Sheet1', 0, colume_to_add_null)
        # 将 从点击屏幕开流 到 p2p唤醒成功 的时间,存储在 ./data.xlsx 路径文件中的 'E', 'F', 'G' 列最后一行
        colume_to_add_null = ['E', 'F', 'G']
        content_to_add_null = ['p2p连接为空', 'Preview 为空', '休眠为空']
        result_save_excel_full_list('./data.xlsx', 'Sheet1', content_to_add_null, colume_to_add_null)
        # 获取完成后清除logcat
        r_obj = os.popen("adb logcat -c")
        r_obj.close()
    else:
        # 将bytes转换成字符串
        last_line = lines[-1].decode('utf-8')
        print("最后一行：", last_line)
        # 计算  从点击屏幕开流 到 p2p唤醒成功 的时间差值
        time_difference_success = time_difference(last_line, click_time)
        results = last_line.split()
        print(results)
        # 查找 ensureDeviceWake wakeSuccess 字段
        for word in results:
            if 'success' in word:
                dev_p2p_state = word
                # 将 从点击屏幕开流 到 p2p唤醒成功 的时间,存储在 ./data.xlsx 路径文件中的 AM 列最后一行
                colume_to_add_success = ['AM']
                result_save_excel_column_list('./data.xlsx', 'Sheet1', 1, colume_to_add_success)
                # 将 从点击屏幕开流 到 p2p唤醒成功 的时间,存储在 ./data.xlsx 路径文件中的 E  列最后一行
                colume_to_add_success = ['E']
                result_save_excel_column_list('./data.xlsx', 'Sheet1', time_difference_success, colume_to_add_success)
                break
        else:
            dev_p2p_state = 'p2pState:failed'  # 如果最后一行中,没有success即p2p连接为失败, dev_p2p_state 赋值为失败状态.
            # 将 从点击屏幕开流 到 p2p唤醒成功 的次数,存储在 ./data.xlsx 路径文件中的 'AM', 'AN', 'AO' 列最后一行
            colume_to_add_failed = ['AM', 'AN', 'AO']
            result_save_excel_column_list('./data.xlsx', 'Sheet1', 0, colume_to_add_failed)
            # 将 从点击屏幕开流 到 p2p唤醒成功 的时间,存储在 ./data.xlsx 路径文件中的 'E', 'F', 'G' 列最后一行
            colume_to_add_failed = ['E', 'F', 'G']
            content_to_add_failed = ['p2p连接失败', 'Preview 失败', '休眠失败']
            result_save_excel_full_list('./data.xlsx', 'Sheet1', content_to_add_failed, colume_to_add_failed)
            # 获取完成后清除logcat
            r_obj = os.popen("adb logcat -c")
            r_obj.close()

    return dev_p2p_state


def get_dev_play_state_result(dev_id, click_time):
    cmd = 'adb -s %s shell logcat -v time -s LivePlayer "\| grep -e state"' % dev_id
    print(cmd)

    p_obj = subprocess.Popen(cmd, shell=False, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
    # 等待日志获取完成
    time.sleep(4)
    # 获取完成后药结束子进程
    p_obj.terminate()
    p_obj.kill()
    # 获取结束后读取内容
    lines = p_obj.stdout.readlines()

    # 如果获取到的内容为空,执行返回值状态为空,否则,根据返回内容,返回成功或失败
    if len(lines) == 0:
        play_state = 'state:null'  # 如果最后一行中,没有 Playing 即 第一帧回调失败, play_state 赋值为失败状态.
        # 将从点击屏幕开流到设备第一帧画面成功回来的时间,存储在 ./data.xlsx 路径文件中的 'AN', 'AO' 列最后一行
        colume_to_add_null = ['AN', 'AO']
        result_save_excel_column_list('./data.xlsx', 'Sheet1', 0, colume_to_add_null)
        # 将从点击屏幕开流到设备第一帧画面成功回来的时间,存储在 ./data.xlsx 路径文件中的 'F', 'G' 列最后一行
        colume_to_add_null = ['F', 'G']
        content_to_add_null = ['Preview 为空', '休眠为空']
        result_save_excel_full_list('./data.xlsx', 'Sheet1', content_to_add_null, colume_to_add_null)
        # 获取完成后清除logcat
        r_obj = os.popen("adb logcat -c")
        r_obj.close()
    else:
        # 将bytes转换成字符串
        last_line = lines[-1].decode('utf-8')
        print("最后一行：", last_line)
        results = last_line.split()

        # 计算 从点击屏幕到唤醒成功的时间差值
        time_difference_success = time_difference(last_line, click_time)

        # 查找playState字段
        for word in results:
            if 'state:Playing' in word:
                play_state = word
                # 将从点击屏幕开流到设备唤醒成功的时间,存储在 ./data.xlsx 路径文件中的 AN 列最后一行
                colume_to_add_success = ['AN']
                result_save_excel_column_list('./data.xlsx', 'Sheet1', 1, colume_to_add_success)
                # 将从点击屏幕开流到设备唤醒成功的时间,存储在 ./data.xlsx 路径文件中的 F 列最后一行
                colume_to_add_success = ['F']
                result_save_excel_column_list('./data.xlsx', 'Sheet1', time_difference_success, colume_to_add_success)
                break
        else:
            play_state = 'state:failed'  # 如果最后一行中,没有 Playing 即 第一帧回调失败, play_state 赋值为失败状态.
            # 将从点击屏幕开流到设备第一帧画面成功回来的时间,存储在 ./data.xlsx 路径文件中的 'AN', 'AO' 列最后一行
            colume_to_add_failed = ['AN', 'AO']
            result_save_excel('./data.xlsx', 'Sheet1', 0, colume_to_add_failed)
            # 将从点击屏幕开流到设备第一帧画面成功回来的时间,存储在 ./data.xlsx 路径文件中的 'F', 'G' 列最后一行
            colume_to_add_failed = ['F', 'G']
            content_to_add_failed = ['Preview 失败', '休眠失败']
            result_save_excel('./data.xlsx', 'Sheet1', content_to_add_failed, colume_to_add_failed)
            # 获取完成后清除logcat
            r_obj = os.popen("adb logcat -c")
            r_obj.close()

    # 获取完成后清除logcat
    r_obj = os.popen("adb logcat -c")
    r_obj.close()

    return play_state


def get_dev_dormancy_state_result(dev_id, click_time):
    # adb shell logcat - v time - s TAG名 （-s表示只抓取这个TAG的日志）
    dp = '\'\{\\"149\\":false\}\''
    cmd = 'adb -s %s shell logcat -v time -s dp_TyDeviceCenter "\| grep -e %s"' % (dev_id, dp)
    print(cmd)

    p_obj = subprocess.Popen(cmd, shell=False, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
    # 等待日志获取完成
    time.sleep(4)
    # 获取完成后药结束子进程
    p_obj.terminate()
    p_obj.kill()
    # 获取结束后读取内容
    lines = p_obj.stdout.readlines()
    print(lines)
    # 如果获取到的内容为空,执行返回值状态为空,否则,根据返回内容,返回成功或失败
    if len(lines) == 0:
        dev_dormancy_state = 'dpStr={"149":null}'  # 如果最后一行中,没有 '\"149\":false' 即 设备依然在唤醒状态 .
        # 将 从点击屏幕返回 到 设备休眠成功 的次数,存储在 ./data.xlsx 路径文件中的 'AO' 列最后一行
        colume_to_add_null = ['AO']
        result_save_excel_column_list('./data.xlsx', 'Sheet1', 0, colume_to_add_null)
        # 将 从点击屏幕返回 到 设备休眠成功 的时间,存储在 ./data.xlsx 路径文件中的 G 列最后一行
        colume_to_add_null = ['G']
        result_save_excel_column_list('./data.xlsx', 'Sheet1', '休眠状态为空', colume_to_add_null)
        # 获取完成后清除logcat
        r_obj = os.popen("adb logcat -c")
        r_obj.close()
    else:
        # 将bytes转换成字符串
        last_line = lines[-1].decode('utf-8')
        print("最后一行：", last_line)

        # 计算  从点击屏幕返回 到 设备休眠成功 的时间差值
        time_difference_success = time_difference(last_line, click_time)

        results = last_line.split()
        print(results)
        # 查找 ensureDeviceWake wakeSuccess 字段
        for word in results:
            if 'dpStr={"149":false}' in word:
                dev_dormancy_state = word
                # 将 从点击屏幕返回 到 设备休眠成功 的时间,存储在 ./data.xlsx 路径文件中的 'AO' 列最后一行
                colume_to_add_success = ['AO']
                result_save_excel_column_list('./data.xlsx', 'Sheet1', 1, colume_to_add_success)
                # 将 从点击屏幕返回 到 设备休眠成功 的时间,存储在 ./data.xlsx 路径文件中的 G 列最后一行
                colume_to_add_success = ['G']
                result_save_excel_column_list('./data.xlsx', 'Sheet1', time_difference_success, colume_to_add_success)
                break
        else:
            dev_dormancy_state = 'dpStr={"149":true}'  # 如果最后一行中,没有 '\"149\":false' 即 设备依然在唤醒状态 .
            # 将 从点击屏幕返回 到 设备休眠成功 的时间,存储在 ./data.xlsx 路径文件中的 'AO' 列最后一行
            colume_to_add_failed = ['AO']
            result_save_excel_column_list('./data.xlsx', 'Sheet1', 0, colume_to_add_failed)
            # 将 从点击屏幕返回 到 设备休眠成功 的时间,存储在 ./data.xlsx 路径文件中的 G 列最后一行
            colume_to_add_failed = ['G']
            result_save_excel_column_list('./data.xlsx', 'Sheet1', '休眠失败', colume_to_add_failed)
            # 获取完成后清除logcat
            r_obj = os.popen("adb logcat -c")
            r_obj.close()

    # 获取完成后清除logcat
    r_obj = os.popen("adb logcat -c")
    r_obj.close()

    return dev_dormancy_state


def get_android_version():
    dev_id = get_dev_id()

    cmd = 'adb -s ' + dev_id + ' shell getprop ro.build.version.release'
    with os.popen(cmd, 'r') as f_obj:
        lines = f_obj.readlines()
        dev_android_ver = re.findall(r'^\w*\b', lines[0])[0]
    return dev_android_ver


def get_package_name():
    cmd = 'aapt dump badging ' + './resource/aosu_app_android_internal_global_debug_3.4.22.8774_1733908106479.apk'
    with os.popen(cmd, 'r') as f_obj:
        lines = f_obj.readlines()
        app_package = re.findall(r'\'com\w*.*?\'', lines[0])[0]
        package_name = app_package[1:-1]

    return package_name


def get_app_version_name():
    """
    ：desc: 获取app的版本号，用于校验关于页面中的版本号
    :return: 返回版本号
    """
    cmd = 'aapt dump badging ' + './resource/aosu_app_android_internal_global_debug_3.0.18.7445_1711101380457.apk'
    with os.popen(cmd, 'r') as f_obj:
        lines = f_obj.readlines()
        app_version_name = re.findall(r'\d+\.\d+\.\d+\.[0-9]{4}', lines[0])[0]

    return app_version_name


def isAppExist():
    dev_id = get_dev_id()
    package_name = get_package_name()

    cmd = 'adb -s ' + dev_id + ' shell pm list packages'
    with os.popen(cmd, 'r') as f_obj:
        exist_packages = f_obj.readlines()

    # 去掉前面的package:和后面的\n
    transform = []

    for i in exist_packages:
        transform.append(i.split(':')[1][:-1])

    if package_name in transform:
        return True
    else:
        return False


def uninstallApp():
    dev_id = get_dev_id()
    package_name = get_package_name()

    if isAppExist():
        # adb install 和 uninstall执行完成后都会返回Success
        os.system('adb -s ' + dev_id + ' uninstall ' + package_name)
        time.sleep(5)


def installApp():
    dev_id = get_dev_id()
    os.system(
        'adb -s ' + dev_id + ' install ' + './resource/aosu_app_android_internal_global_debug_3.0.18.7445_1711101380457.apk')
    time.sleep(5)


def isAwake():
    dev_id = get_dev_id()
    cmd = 'adb -s ' + dev_id + ' shell dumpsys window policy'
    screen_awake_value = '      screenState=SCREEN_STATE_ON\n'

    with os.popen(cmd, 'r') as f_obj:
        all_list = f_obj.readlines()

    if screen_awake_value in all_list:
        return True
    else:
        return False


def setScreenAlwaysOn():
    dev_id = get_dev_id()
    # 设置亮屏时间时永不锁屏 -1
    os.system('adb -s ' + dev_id + ' shell settings put system screen_off_timeout -1')

    # def initPhone(): os.system('adb shell settings put system screen_off_timeout 600000') 设置默认壁纸IDLE.png os.system(
    # 'adb push ../resource/IDLE.png /sdcard/Download/') os.system('adb shell am start -d
    # file:////sdcard/Download/IDLE.png -a android.service.wallpaper.CROP_AND_SET_WALLPAPER -f 0x1
    # com.android.launcher3/.WallpaperCropActivity') os.system('adb shell am start -a
    # android.intent.action.ATTACH_DATA -c android.intent.category.DEFAULT -d file://sdcard/Download/IDLE.png')


def keyEventSend(keycode):
    dev_id = get_dev_id()
    cmd = 'adb -s %s shell input keyevent %d' % (dev_id, keycode)
    os.system(cmd)
